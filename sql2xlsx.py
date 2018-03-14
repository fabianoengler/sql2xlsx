#!/bin/env python3
"""
Copyright 2018 (C) Fabiano Engler Neto - fabianoengler(at)gmail(.)com

Simple script and class helper to execute a MySQL query and automatically
produce a nice formated XLSX output file.
"""

import sys
import mysql.connector
import os
from openpyxl import Workbook
from openpyxl.utils.cell import get_column_letter
from openpyxl.styles import Font
from openpyxl import load_workbook
from openpyxl.styles import Alignment
from decimal import Decimal
from math import ceil
from collections import Counter
from tempfile import NamedTemporaryFile
import logging

log = logging.getLogger(__name__)
log.addHandler(logging.NullHandler())

verbosity_level = 2
#verbosity_level = 5


def verb(level, *args, **kwargs):
    kwargs['flush'] = True
    if level <= verbosity_level:
        print(*args, **kwargs)


class MySql2Xlsx(object):

    MIN_COL_WIDTH = 6
    MIN_COL_FULL_LEN = 25
    NUMBER_FORMAT = '#,##0.00'
    FETCH_CHUNK_SIZE = 1000

    def __init__(self, mysql_config=None, query_str=None, out_fname=None,
                 query_params=None):
        self.__mysql_config = mysql_config
        self.query_str = query_str 
        self.tmp_fname = self.out_fname = out_fname
        self.query_params = query_params
        self.conn = None
        self.cursor = None

    def generate_report(self):
        self.mysql_connect()
        self.mysql_execute()
        self.create_workbook()
        self.prepare_sheet()
        self.fetch_rows_and_write()
        self.mysql_disconnect()
        self.make_final_adjustments()
        self.write_final_file()
        verb(1, 'Done.')

    def mysql_connect(self, mysql_config=None):
        verb(1, 'Connecting...')
        if mysql_config is None:
            mysql_config = self.__mysql_config 
        if mysql_config is None:
            raise ValueError('No mysql_config defined')

        self.conn = mysql.connector.connect(**mysql_config)
        self.cursor = self.conn.cursor()
        return self.cursor

    def mysql_execute(self, query_str=None, query_params=None):
        verb(1, 'Executing SQL...')
        if query_str is None:
            query_str = self.query_str
        if query_str is None:
            raise ValueError('No query_str defined')
        if query_params is None:
            query_params = self.query_params

        self.cursor.execute(query_str, query_params)
        if not len(self.cursor.column_names):
            raise RuntimeError('No columns on query set')
        return self.cursor

    def mysql_disconnect(self):
        if self.conn is not None:
            verb(3, 'Disconnecting...')
            self.conn.close()
            self.conn = None

    def create_workbook(self):
        verb(1, 'Creating XLSX Workbook...')
        self.wb = Workbook(write_only=True)
        return self.wb

    def create_worksheet(self):
        verb(3, 'Creating Worksheet...')
        self.ws = self.wb.create_sheet()
        return self.ws

    def freeze_panes(self):
        verb(3, 'Freezing first row...')
        self.ws.freeze_panes = 'A2'
        return self.ws

    def set_filters(self):
        verb(3, 'Setting Auto Filter...')
        self.ws.auto_filter.ref = 'A:{}'.format(
                get_column_letter(len(self.cursor.column_names)))
        return self.ws

    def prepare_sheet(self):
        self.create_worksheet()
        self.freeze_panes()
        self.set_filters()
        self.write_column_names()
        return self.ws

    def mysql_fetch_data_chunk(self, chunk_size=FETCH_CHUNK_SIZE):
        return self.cursor.fetchmany(size=chunk_size)

    def mysql_fetch_chunk_iterator(self, chunk_size=FETCH_CHUNK_SIZE):
        while True:
            data_chunk = self.mysql_fetch_data_chunk(chunk_size)
            if not data_chunk:
                return
            yield data_chunk

    def mysql_fetch_rows_iterator(self, chunk_size=FETCH_CHUNK_SIZE):
        for data_chunk in self.mysql_fetch_chunk_iterator(chunk_size):
            verb(2, '.', end='')
            for data_row in data_chunk:
                yield data_row
        verb(2, '')

    def write_column_names(self):
        verb(3, 'Writing column names on first row...')
        sheet_row = []
        for col in self.cursor.column_names:
            col_name = col.decode('utf-8') if isinstance(col, bytes) else col
            val = col_name.replace('_', ' ').title()
            sheet_row.append(val)
        self.ws.append(sheet_row)
        return self.ws

    def _fetch_write_loop_start(self):
        cols_lengths = [ [] for _ in range(len(self.cursor.column_names)) ]
        cols_types = [ Counter() for _ in range(len(self.cursor.column_names)) ]

        self.cols_lengths = cols_lengths 
        self.cols_types = cols_types 

    def _fetch_write_loop_step(self, data_row):
        cols_lengths = self.cols_lengths 
        cols_types = self.cols_types 
        sheet_row = []
        for i, value in enumerate(data_row):
            sheet_row.append(value)
            chars = 0 if value is None else len(str(value))
            if isinstance(value, int):
                chars += 1
            elif isinstance(value, (float, Decimal)):
                chars += 2
            cols_types[i][type(value)] += 1
            cols_lengths[i].append(0 if value is None else len(str(value)))
        self.ws.append(sheet_row)

    def _fetch_write_loop_finish(self):
        pass

    def fetch_rows_and_write(self, chunk_size=FETCH_CHUNK_SIZE):

        verb(1, 'Writing data to worksheet...')

        self._fetch_write_loop_start()
        for data_row in self.mysql_fetch_rows_iterator(chunk_size):
            self._fetch_write_loop_step(data_row)

        verb(2, 'Total Number of rows: {}'.format(self.cursor.rowcount))
        self._fetch_write_loop_finish()

        return self.ws

    def _check_tmp_fname(self):
        log.debug('Checking tmp file name...')
        if self.tmp_fname is None:
            f = NamedTemporaryFile(delete=False, suffix='.xlsx')
            self.tmp_fname = f.name
            f.close()
            log.debug('Tmp file name: "%s"', self.tmp_fname)

    def _cleanup_tmp_file(self):
        log.debug('Cleaning up tmp file...')
        if self.tmp_fname is not None and self.tmp_fname != self.out_fname :
            log.debug('Unlinking file: "%s"...', self.tmp_fname)
            os.unlink(self.tmp_fname)
            self.tmp_fname = None

    def save_and_reload(self):
        self._check_tmp_fname()
        verb(3, 'Saving intermediate file...')
        self.wb.save(self.tmp_fname)
        verb(3, 'Reloading intermediate file...')
        self.wb = load_workbook(self.tmp_fname)
        self.ws = self.wb.active

    def make_final_adjustments(self):
        verb(1, 'Final adjustments...')

        # need to save and re-open file to edit, as write_only mode does not
        # support in-memory editing
        self.save_and_reload()

        self.resize_columns()
        self.format_numbers()
        self.format_column_names()

        return self.ws

    def resize_columns(self):
        verb(2, 'Resizing columns...')
        for i, col in enumerate(self.cols_lengths):
            m = max(col)
            if m <= self.MIN_COL_FULL_LEN:
                column_width = max(m, self.MIN_COL_WIDTH)
            else:
                dec9th = sorted(col)[ceil(len(col)/10*9)]
                column_width = min(m, dec9th)
            log.debug('column_width : %s', column_width)

            column = get_column_letter(i+1)
            self.ws.column_dimensions[column].width = column_width

    def format_numbers(self):
        verb(2, 'Formating numbers...')
        for i, cells_types in enumerate(self.cols_types):
            del cells_types[type(None)]
            try:
                col_type = cells_types.most_common(1)[0][0]
            except IndexError:
                continue  # no other types besides None

            if col_type in (float, Decimal):
                for cell in self.ws[get_column_letter(i+1)]:
                    cell.number_format = self.NUMBER_FORMAT

    def format_column_names(self, font=None, alignment=None):
        verb(2, 'Formating column names...')
        if alignment is None:
            alignment = Alignment(wrap_text=True)
        if font is None:
            font = Font(bold=True)

        for i in range(len(self.cursor.column_names)):
            cell = self.ws.cell(row=1, column=i+1)
            cell.alignment = alignment 
            cell.font = font

        row1 = self.ws.row_dimensions[1]
        row1.height = 26

    def write_final_file(self, out_fname=None):
        if out_fname is None:
            out_fname = self.out_fname
        if out_fname is None:
            raise ValueError('No output file name defined')
        verb(1, 'Writing final file...')
        self.wb.save(out_fname)

    def __del__(self):
        self.mysql_disconnect()
        self._cleanup_tmp_file()


def main():
    # logging.basicConfig(level=logging.DEBUG)
    logging.basicConfig()

    if not (2 <= len(sys.argv) <= 3) or sys.argv[1] in ('-h', '--help'):
        print('Usage:')
        print('    {} <query-file.sql> [output_file.xlsx]'.format(sys.argv[0]))
        sys.exit(-1)

    try:
        raw_sql = open(sys.argv[1], 'rt').read()
    except OSError as e:
        print(e)
        sys.exit(e.errno)

    from config import mysql_config 

    if len(sys.argv) == 3:
        out_fname = sys.argv[2]
    else:
        out_fname = '{}_result.xlsx'.format(sys.argv[1])

    mysql2xlsx = MySql2Xlsx(mysql_config, raw_sql, out_fname)
    mysql2xlsx.generate_report()


if __name__ == '__main__':
    main()

